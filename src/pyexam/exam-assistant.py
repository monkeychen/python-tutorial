import os
import argparse
import shutil
from datetime import datetime
from openpyxl import load_workbook
from unrar import rarfile
import tkinter as tk


FMT_DEFAULT = "%Y-%m-%d %H:%M:%S"


class Answer(object):
    score_conf = [
        {"title": "single_choice", "unit": 2},
        {"title": "multi_choice", "unit": 3},
        {"title": "judgment", "unit": 2},
        {"title": "programme_fill", "unit": 5},
        {"title": "complex", "unit": 10},
    ]

    def __init__(self, single_choice=None, multi_choice=None, judgment=None):
        self.single_choice = single_choice
        self.multi_choice = multi_choice
        self.judgment = judgment
        self.programme_fill = []
        self.complex = []


class Student(object):

    def __init__(self, name, answer: Answer = None, std_answer: Answer = None, valid=True):
        self.name = name
        self.answer = answer
        self.score = 0
        self.std_answer = std_answer
        self.valid = valid
        self.answer_dir_path = None

    def calc_score(self):
        if self.answer is None:
            self.valid = False
            return
        for conf in Answer.score_conf:
            unit = conf.get("unit")
            q_type = conf.get("title")
            tmp_std_ans_values = getattr(self.std_answer, q_type)
            tmp_ans_values = getattr(self.answer, q_type)
            if len(tmp_ans_values) != len(tmp_std_ans_values):
                self.valid = False
                return
            for i in range(0, len(tmp_std_ans_values)):
                self.score = self.score + (unit if tmp_std_ans_values[i] == tmp_ans_values[i] else 0)
        self.valid = self.score < 20
        return self.score


class ExamAssistant(object):
    root_dir = os.getcwd()
    question_conf = [
        {"begin": "程序填空题01", "end": "txt"},
        {"begin": "程序填空题02", "end": "txt"},
        {"begin": "程序填空题03", "end": "txt"},
        {"begin": "综合题01", "end": "txt"},
        {"begin": "综合题02", "end": "txt"}
    ]

    def __init__(self, tk_inst: tk.Tk):
        self.tk_inst = tk_inst
        self.args = self.parse_cli_args()
        self.root_dir = self.args.root_dir
        self.answer_file_num = self.args.answer_file_num
        self.answer_dir = f"{self.root_dir}/answer"
        self.upload_dir = f"{self.root_dir}/upload"
        self.temp_dir = f"{self.upload_dir}/temp"
        self.action = self.args.action
        self.std_answer = self.load_answer_file(f"{self.answer_dir}/答题卡.xlsx")
        self.student_list = []
        self.init_win_layout()

    def init_win_layout(self):
        self.tk_inst.title("压缩包检测工具V1.0")
        self.tk_inst.geometry('900x600')

        self.btn = tk.Button(self.tk_inst, text="压缩包内容检测", bg="lightblue", width=30, height=1, command=self.execute)
        self.btn.pack(pady=15)
        self.label3 = tk.Label(self.tk_inst, text="本产品只获取当前文件夹中rar压缩包的总数，\n以及检测rar中是否包含Acopr与prog、答题卡xlsx文件。\n如需更详细的检测，请进行人工判定")
        self.label3.pack(pady=1)
        self.text2 = tk.StringVar()
        self.label2 = tk.Label(self.tk_inst, textvariable=self.text2, fg="red", width=90, relief="sunken", bg="white", font=16)
        self.label2.pack(pady=1)

        self.text1 = tk.StringVar()
        self.label1 = tk.Label(self.tk_inst, textvariable=self.text1, fg="red", width=90, height=30, relief="sunken", bg="white", font=16)
        self.label1.pack()

    def load_answer_file(self, answer_file_path):
        answer_card_wb = load_workbook(answer_file_path)
        sheet = answer_card_wb["Sheet1"]
        parse_conf = {
            "single_choice": {'row': 3, 'max_col': 16},
            "multi_choice": {'row': 6, 'max_col': 6},
            "judgment": {'row': 9, 'max_col': 16}
        }
        answer = Answer([], [], [])
        for k in parse_conf.keys():
            conf = parse_conf.get(k)
            if conf is None:
                continue
            tmp_arr = []
            for col_idx in range(2, conf.get('max_col') + 1):
                tmp_arr.append(sheet.cell(conf.get('row'), col_idx).value)
            setattr(answer, k, tmp_arr)
        return answer

    def submit_check(self):
        self._remove_dir_tree(dir_path=self.temp_dir, include_self=False)
        dir_list = os.listdir(self.upload_dir)
        check_report = []
        student_list = []
        for stu_main_dir_name in dir_list:
            stu_main_dir_path = f"{self.upload_dir}/{stu_main_dir_name}"
            if "temp" == stu_main_dir_name or "py-code" == stu_main_dir_name or os.path.isfile(stu_main_dir_path) or len(os.listdir(stu_main_dir_path)) == 0:
                print(f"{stu_main_dir_path} is file or is an empty directory or a temp dir!")
                continue
            student = Student(name=stu_main_dir_name, std_answer=self.std_answer)
            student_list.append(student)
            rar_cnt = 0
            rar_file_path = None
            for stu_file_name in os.listdir(stu_main_dir_path):
                if not stu_file_name.endswith("rar"):
                    continue
                rar_cnt += 1
                if rar_cnt > 1:
                    check_report.append(f"目录【{stu_main_dir_name}]下发现多份RAR文件，请人工确认是否异常！")
                    student.valid = False
                    break
                rar_file_path = f"{stu_main_dir_path}/{stu_file_name}"

            if student.valid:
                unrar_tmp_dir_path = f"{self.temp_dir}/{stu_main_dir_name}"
                try:
                    self._remove_dir_tree(dir_path=unrar_tmp_dir_path, include_self=False)
                    rar_file = rarfile.RarFile(rar_file_path)
                    matched_name_list = []
                    card_name = None
                    for name_in_rar in rar_file.namelist():
                        base_name = os.path.basename(name_in_rar)
                        if self._is_matched_file(base_name):
                            matched_name_list.append(name_in_rar)
                        elif base_name == "答题卡.xlsx":
                            card_name = name_in_rar

                    if len(matched_name_list) < self.answer_file_num or card_name is None:
                        check_report.append(f"压缩文件【{rar_file_path}]下的文件数量不满意要求，请人工确认！")
                        student.valid = False
                        continue
                    if student.valid:
                        rar_file.extractall(path=unrar_tmp_dir_path)
                        card_path = f"{unrar_tmp_dir_path}/{card_name}"
                        student.answer_dir_path = os.path.dirname(card_path)
                        if os.path.exists(card_path):
                            student.answer = self.load_answer_file(card_path)
                            student.calc_score()
                        else:
                            check_report.append(f"无法加载答题卡文件【{card_path}】，请人工确认！")
                            student.valid = False
                except Exception as e:
                    print(f"Fail to parse student[{stu_main_dir_name}]'s answer rar file!")

        invalid_student_list = [stu for stu in student_list if not stu.valid]
        for stu in invalid_student_list:
            if stu.score < 20:
                check_report.append(f"学生【{stu.name}】的成绩为【{stu.score}】，低于20分，请人工确认答卷是否异常！")
        return check_report

    def evaluate(self):
        pass

    def _remove_dir_tree(self, dir_path, include_self=True):
        if os.path.exists(dir_path):
            shutil.rmtree(dir_path)
        if not include_self:
            os.makedirs(dir_path)

    def _list_dir(self, dir_path, ext=None):
        name_list = []
        for dir_name in os.listdir(dir_path):
            if ext is None:
                name_list.append(dir_name)
            else:
                if dir_name.endswith(ext):
                    name_list.append(dir_name)
        return name_list

    def _is_matched_file(self, file_name: str):
        matched = False
        for conf in self.question_conf:
            matched = file_name.startswith(conf.get("begin")) and file_name.endswith(conf.get("end"))
            if matched:
                break
        return matched

    def datetime_to_str(self, dt, fmt=FMT_DEFAULT):
        return dt.strftime(fmt)

    def parse_cli_args(self):
        cli_parser = argparse.ArgumentParser(description="这是一个Python考试辅助工具。", add_help=False)
        cli_parser.add_argument('-a', '--action', nargs='?', choices=['submit', 'evaluate'], default='submit', help='设置运行环境（submit：交卷检测，evaluate：评卷），默认为：交卷检测')
        cli_parser.add_argument('-e', '--env_mode', nargs='?', choices=['dev', 'prod'], default='prod', help='设置运行环境（dev：开发环境，prod：生产环境），默认为：开发环境')
        cli_parser.add_argument('-n', '--answer_file_num', nargs='?', type=int, default=5, help='学生提交的python文件数，默认为5。')
        cli_parser.add_argument('-r', '--root_dir', nargs='?', default=f"{os.getcwd()}", help='程序工作根目录路径。')
        cli_parser.add_argument('-x', '--ext_name', nargs='?', default="py", help='学生提交作业中程序文件扩展名，默认为：py')
        cli_parser.add_argument('-h', '--help', action='help', help='显示本帮助信息并退出')
        cli_parser.add_argument('-v', '--version', action='version', version='%(prog)s V0.0.1', help='显示当前版本信息并退出')

        args = cli_parser.parse_args()
        scheduled_time = self.datetime_to_str(datetime.now(), fmt=FMT_DEFAULT)
        print(f"The programme[{cli_parser.prog}] was scheduled at {scheduled_time}, cwd = {os.getcwd()}, env_mode = {args.env_mode}, "
              f"root_dir = {args.root_dir}, answer_file_num = {args.answer_file_num}")
        return args

    def execute(self):
        if self.action == 'submit':
            check_result = self.submit_check()
            infos = "\n".join(check_result)
            self.text1.set(infos)
            for info in check_result:
                print(info)
        else:
            self.evaluate()


class PythonQuestionEval(object):

    question_conf = ExamAssistant.question_conf

    def __init__(self, answer_dir_path=None):
        if answer_dir_path is None:
            answer_dir_path = f"{os.getcwd()}/upload/temp"
        self.answer_dir_path = answer_dir_path

    def eval_question_1(self, question_file_path):
        pass

    def eval_question_2(self, question_file_path):
        pass

    def eval_question_3(self, question_file_path):
        pass

    def eval_question_4(self, question_file_path):
        pass

    def eval_question_5(self, question_file_path):
        pass

    def evaluate(self):
        for stu_main_dir_name in os.listdir(self.answer_dir_path):
            stu_main_dir_path = f"{self.answer_dir_path}/{stu_main_dir_name}"



if __name__ == "__main__":
    win = tk.Tk()
    GUI = ExamAssistant(win)
    win.mainloop()
